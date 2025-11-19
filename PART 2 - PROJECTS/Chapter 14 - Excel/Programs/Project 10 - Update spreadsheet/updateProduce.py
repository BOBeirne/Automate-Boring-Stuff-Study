# Corrects costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales3.xlsx')
sheet = wb['Sheet']

# Dictionary with updated data
PRICE_UPDATES = {'Garlic': 3.07,
					'Celery': 1.19,
					'Lemon': 1.27}

# TODO loop through rows and update prices for matching keys
for row_num in range(2, sheet.max_row +1):
	produce_name = sheet.cell(row=row_num, column=1).value
	if produce_name in PRICE_UPDATES:
		sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]
		print('Data updated.')
print('Saving data to new file')
wb.save('updatedProduceSales3.xlsx')
print('done')