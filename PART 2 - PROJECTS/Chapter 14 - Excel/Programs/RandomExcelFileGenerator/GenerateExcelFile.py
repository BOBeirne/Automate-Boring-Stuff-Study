# This program will generate an Excel file with random data

import openpyxl, random

wb = openpyxl.Workbook()
sheet = wb.active


fileName = input('What will be the name of the file? ')
sheetNr = input('How many sheets?')
sheetName = ('Provide Sheet {sheetNr} name: '.format(sheetNr=sheetNr))
colLen = input('How many columns? ')
rowLen = input('How many rows? ')

if int(sheetNr) >= 1:
	for i in range(1, int(sheetNr)):
		sheet = wb.create_sheet(title=input(sheetName))
		for row in range(1, int(rowLen)):
			for col in range(1, int(colLen)):
				cell = sheet.cell(row=row, column=col)
				cell.value = random.randint(1, 100)
wb.save(fileName + '.xlsx')
print(f'File {fileName} created.')